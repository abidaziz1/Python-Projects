# • Read data from an Excel spreadsheet.
# • Find all members who have not paid dues for the latest month.
# • Find their email addresses and send them personalized reminders.
# This means your code will need to do the following:
# • Open and read the cells of an Excel document with the openpyxl mod-
# ule. (See Chapter 12 for working with Excel files.)
# • Create a dictionary of members who are behind on their dues.
# • Log in to an SMTP server by calling smtplib.SMTP(), ehlo(), starttls(),
# and login().
# • For all members behind on their dues, send a personalized reminder
# email by calling the sendmail() method.

#Step 1: Open the Excel File
import openpyxl, smtplib, sys
# Open the spreadsheet and get the latest dues status.
wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
lastCol = sheet.get_highest_column()
latestMonth = sheet.cell(row=1, column=lastCol).value

# After importing the openpyxl, smtplib, and sys modules, we open our
# duesRecords.xlsx file and store the resulting Workbook object in wb . Then we
# get Sheet 1 and store the resulting Worksheet object in sheet . Now that we
# have a Worksheet object, we can access rows, columns, and cells. We store the
# highest column in lastCol , and we then use row number 1 and lastCol to
# access the cell that should hold the most recent month. We get the value in
# this cell and store it in latestMonth 


# Step 2: Find All Unpaid Members
unpaidMembers = {}
for r in range(2, sheet.get_highest_row() + 1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email
# This code sets up an empty dictionary unpaidMembers and then loops
# through all the rows after the first. For each row, the value in the most
# recent column is stored in payment. If payment is not equal to 'paid', then
# the value of the first column is stored in name , the value of the second col-
# umn is stored in email , and name and email are added to unpaidMembers.


# Step 3: Send Customized Email Reminders
smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login('dipumoniminister@gmail.com', sys.argv[1]) # sys.argv[1], which will store your password string. You’ll enter the password as a command line argument each time you run the program, 

# Send out reminder emails.
for name, email in unpaidMembers.items():
    body = "Subject: %s dues unpaid.\nDear %s,\nRecords show that you have not
paid dues for %s. Please make this payment as soon as possible. Thank you!'" %
(latestMonth, name, latestMonth)
    print('Sending email to %s...' %email)
    sendmailStatus = smtpObj.sendmail('dipumoniminister@gmail.com', email, body)
    if sendmailStatus != {}:
        print('There was a problem sending email to %s: %s' % (email,
        sendmailStatus))
smtpObj.quit()

# each member who hasn’t paid, we customize a message with the latest month
# and the member’s name, and store the message in body. We print output
# saying that we’re sending an email to this member’s email address. Then
# we call sendmail(), passing it the from address and the customized message.
# We store the return value in sendmailStatus.
# Remember that the sendmail() method will return a nonempty diction-
# ary value if the SMTP server reported an error sending that particular
# email. The last part of the for loop at checks if the returned dictionary is
# nonempty, and if it is, prints the recipient’s email address and the returned
# dictionary.
# After the program is done sending all the emails, the quit() method is
# called to disconnect from the SMTP server.


